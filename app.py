from flask import Flask, render_template, request, redirect, url_for,jsonify
from flask import send_file
from werkzeug.utils import secure_filename, safe_join
<<<<<<< HEAD
import csv
import os
from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
import smtplib
=======
import os
import smtplib
from datetime import datetime, timedelta
>>>>>>> main
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import chardet
from base import lojas
from materials import materials
from flask_socketio import SocketIO, emit
<<<<<<< HEAD

=======
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from configs import EMAIL_SETTINGS
>>>>>>> main


app = Flask(__name__)
socketio = SocketIO(app)


AUTHORIZED_MATERIALS = []  
status = "" 



<<<<<<< HEAD

@app.route('/config', methods=['GET'])
def configurations():
   
    return render_template('config.html', authorized_materials=', '.join(AUTHORIZED_MATERIALS), status=status)

@app.route('/check-password', methods=['POST'])
def check_password():
    password = request.form.get('password')
    if password == '2024': 
        return redirect(url_for('configurations')) 
    else:
        return redirect(url_for('index', message='Senha incorreta!'))

=======
# ROTA DE CONFIGURAÇÕES: esta rota exibe a página de configurações onde você pode atualizar os materiais autorizados e o status.
@app.route('/config', methods=['GET'])
def configurations():
    # Renderiza o template 'config.html', passando os materiais autorizados e o status atual.
    return render_template('config.html', authorized_materials=', '.join(AUTHORIZED_MATERIALS), status=status)


# FUNÇÃO DE VERIFICAR SENHA: esta rota é usada para verificar a senha inserida no formulário da página de configurações.
@app.route('/check-password', methods=['POST'])
def check_password():
    password = request.form.get('password')  # Obtém a senha do formulário
    if password == '2024':  # Verifica se a senha é correta
        return redirect(url_for('configurations'))  # Redireciona para a página de configurações se a senha estiver correta
    else:
        return redirect(url_for('index', message='Senha incorreta!'))  # Redireciona para a página inicial com uma mensagem de erro


# FUNÇÃO DE ATUALIZAR CONFIGURAÇÕES: esta rota atualiza os materiais autorizados e o status com base nos dados definidos na pagina de configurações.
>>>>>>> main
@app.route('/update-settings', methods=['POST'])
def update_settings():
    global AUTHORIZED_MATERIALS
    global status
<<<<<<< HEAD
    authorized_materials = request.form.get('authorized_materials').split(',')
    status = request.form.get('status')

    AUTHORIZED_MATERIALS = [material.strip() for material in authorized_materials]
    status = status
    socketio.emit('update_notification', {'message': 'Configurações atualizadas!'})
    return redirect(url_for('configurations'))

=======
    # Recebe os materiais autorizados e o status do formulário e atualiza as variáveis globais
    authorized_materials = request.form.get('authorized_materials').split(',')
    status = request.form.get('status')

    # Atualiza a lista de materiais autorizados removendo espaços extras e atualiza o status
    AUTHORIZED_MATERIALS = [material.strip() for material in authorized_materials]
    status = status
    # Emite uma notificação via SocketIO para informar que as configurações foram atualizadas
    socketio.emit('update_notification', {'message': 'Configurações atualizadas!'})
    return redirect(url_for('configurations'))  # Redireciona para a página de configurações


# FUNÇÃO DE LIMPAR CONFIGURAÇÕES: esta rota limpa todas as configurações de materiais autorizados e status.
>>>>>>> main
@app.route('/clear-settings', methods=['POST'])
def clear_settings():
    global AUTHORIZED_MATERIALS
    global status
<<<<<<< HEAD
    AUTHORIZED_MATERIALS = []  
    status = ""  
    return redirect(url_for('configurations'))  # Mudança feita aqui


def send_email_with_attachment(send_to, subject, body, file_path):

    #ALTERAR
    msg = MIMEMultipart()
    msg['From'] = 'contato.paulooliver9@outlook.com'
    msg['To'] = send_to
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

   
=======
    # Limpa as variáveis globais, removendo todos os materiais autorizados e redefinindo o status
    AUTHORIZED_MATERIALS = []  
    status = ""  
    return redirect(url_for('configurations'))  # Redireciona para a página de configurações


# FUNÇÃO DE ENVIO DE E-MAIL: esta função é responsável por enviar um e-mail com um anexo.
def send_email_with_attachment(send_to, subject, body, file_path):
    # Configuração das informações de e-mail obtidas do arquivo 'configs.py'
    email_user = EMAIL_SETTINGS['EMAIL_USER']
    email_password = EMAIL_SETTINGS['EMAIL_PASSWORD']
    smtp_server = EMAIL_SETTINGS['SMTP_SERVER']
    smtp_port = EMAIL_SETTINGS['SMTP_PORT']

    # Prepara a mensagem de e-mail
    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = send_to
    msg['Subject'] = subject

    # Anexa o corpo do e-mail
    msg.attach(MIMEText(body, 'plain'))

    # Anexa o arquivo ao e-mail
>>>>>>> main
    with open(file_path, "rb") as attachment:
        part = MIMEApplication(
            attachment.read(),
            Name=os.path.basename(file_path)
        )
<<<<<<< HEAD
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        msg.attach(part)

  
    server = smtplib.SMTP('smtp.office365.com', 587)
    server.starttls()
   
    server.login(msg['From'], 'Thivi8090p')  
    text = msg.as_string()
    server.sendmail(msg['From'], msg['To'], text)
    server.quit()



def get_encoding(file_path):
    with open(file_path, 'rb') as file:
        return chardet.detect(file.read())['encoding']

    

def update_excel(store, material, description, quantity):
   
    source_filename = 'devolucao.xlsx'
    
   
    wb = load_workbook(source_filename)
    ws = wb.active
    
   
    ws['G7'] = store
    ws['C14'] = material
    ws['D14'] = description
    ws['H14'] = quantity
    ws['L14'] = "Robo"
    
    
    ws.protection = SheetProtection(sheet=True, password='thivi8090#', objects=True, scenarios=True)
    
    
    temp_dir = '/tmp'
    new_filename = os.path.join(temp_dir, f'devolucao_{store}.xlsx')
    wb.save(new_filename)
    
    send_to = 'paulo.cunha@hortifruti.com.br'
    subject = 'Planilha de Devolução'
    body = 'Aqui está a planilha de devolução solicitada.'
    send_email_with_attachment(send_to, subject, body, new_filename)
    
    return new_filename


    
@app.route('/get-description')
def get_description():
    material = request.args.get('material')
    description = ''
    authorized = material in AUTHORIZED_MATERIALS 

    if material:
        material = int(material)  
        description = materials.get(material, '')  

    return jsonify(description=description, authorized=authorized)



def get_stores():
    stores = []

    for store_name in lojas:
        stores.append(store_name)
    
    return stores



@app.route('/', methods=['GET', 'POST'])
def index():
    message = ''
    stores = get_stores() 

    if request.method == 'POST':
        store = request.form.get('store')
        material = request.form.get('material')
        quantity = request.form.get('quantity')
        description = request.form.get('description')

        new_filename = update_excel(store, material, description, quantity) 
        return send_file(new_filename, as_attachment=True, download_name=new_filename)

    return render_template('index.html', message=message, stores=stores, status=status)






if __name__ == '__main__':
    socketio.run(app, debug=True, host="0.0.0.0", port=5000)
=======
        # Define o cabeçalho para o anexo
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        msg.attach(part)

    # Conecta ao servidor SMTP e envia o e-mail
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()  # Inicia a criptografia TLS
    server.login(email_user, email_password)  # Faz o login no servidor SMTP
    text = msg.as_string()  # Converte a mensagem para string
    server.sendmail(email_user, msg['To'], text)  # Envia o e-mail
    server.quit()  # Encerra a conexão com o servidor SMTP



#FUNÇÃO PARA CRIAR O PDF DE AUTORIZAÇÃO DA LOJA
def create_pdf(store, material, description, quantity, authorized_by='Formulário Automatizado', motivo_devolucao='Envio errado'):
    # Configurações iniciais
    project_root_dir = '.'  
    filename = f'devolucao_{store}.pdf'
    file_path = os.path.join(project_root_dir, filename)
    c = canvas.Canvas(file_path, pagesize=letter)
    width, height = letter 


    #Configurar cor de fundo
    background_color = HexColor('#875742')  
    c.setFillColor(background_color)
    c.rect(0, 0, width, height, fill=True, stroke=False)

    # Adicionar marca d'água
    c.setFont("Helvetica-Bold", 60)
    c.setFillAlpha(0.1)  # Torna a marca d'água menos opaca
    c.setFillColorRGB(0, 0, 0)  # Cor da marca d'água: preto
    c.drawCentredString(width / 2.0, height / 2.0, "PROIBIDO ALTERAR")

    
    # Restaurar a opacidade para o restante do texto
    c.setFillAlpha(1)
    c.setFillColorRGB(1, 1, 1) 


    # Adicionar logo
    #logo_path = os.path.join(project_root_dir, 'static/img/logo.png') 
    #c.drawImage(logo_path, 100, height - 100, width=2*inch, height=0.5*inch)

    # Configurar o título e subtitulo
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2.0, height - 120, "AUTORIZAÇÃO DE DEVOLUÇÃO DE MERCADORIA EM EXCEÇÃO")
    c.setFont("Helvetica-Oblique", 12)
    c.drawCentredString(width / 2.0, height - 140, "(Anexar junto à Nota Fiscal)")

    # Configurar informações da autorização
    today = datetime.now()
    max_date = today + timedelta(days=3)  
    c.setFont("Helvetica", 12)
    c.drawString(100, height - 160, f'LOJA/UNIDADE: {store}')
    c.drawString(400, height - 160, f'DATA DA AUTORIZAÇÃO: {today.strftime("%d/%m/%Y")}')
    c.drawString(100, height - 180, f'DATA MÁXIMA PARA ENVIO: {max_date.strftime("%d/%m/%Y")}')

    # Configurar detalhes do produto
    c.drawString(100, height - 220, f'DESCRICÃO DO PRODUTO: {description}')
    c.drawString(100, height - 240, f'QUANTIDADE: {quantity}')
    c.drawString(400, height - 240, f'VALIDADE: {"-"}')  
    c.drawString(100, height - 260, f'LOTE: {"-"}')  #

    # Configurar autorização e motivo da devolução
    c.drawString(100, height - 280, f'DEVOLUÇÃO COM O FÍSICO: {"Sim"}')  
    c.drawString(100, height - 300, f'NOME DE QUEM AUTORIZOU: {authorized_by}')
    c.drawString(100, height - 320, f'QUEM AUTORIZOU CD OU COMERCIAL: {"CD"}')  
    c.drawString(100, height - 340, f'MOTIVO DEVOLUÇÃO: {motivo_devolucao}')

    # Configurar validação do documento
    c.setFont("Helvetica-Oblique", 10)
    footer_start_y_position = 100  
    c.drawString(50, footer_start_y_position, "1° O formulário deve seguir anexo.")
    c.drawString(50, footer_start_y_position - 15, "2° A quantidade informada deverá ser exata ao que constar no formulário.")
    c.drawString(50, footer_start_y_position - 30, "3° A mercadoria só poderá retornar ao CD se estiver íntegra.")

    # Finalizar e salvar o PDF
    c.save()
    print(f"Arquivo PDF gerado: {file_path}")
    return file_path

  
# Esta função retorna a descrição de um material. É um endpoint da API que recebe um ID de material e verifica se ele está na lista de materiais autorizados.
@app.route('/get-description')
def get_description():
    material = request.args.get('material')  # Pega o ID do material dos parâmetros da requisição
    description = ''  # Inicializa a variável de descrição
    authorized = material in AUTHORIZED_MATERIALS  # Verifica se o material está na lista de materiais autorizados

    # Se houver um ID de material fornecido, converte-o para um inteiro e obtém sua descrição do dicionário 'materials', se existir
    if material:
        material = int(material)  # Converte ID do material de string para inteiro
        description = materials.get(material, '')  # Obtém a descrição, retorna uma string vazia se o ID do material não for encontrado

    # Retorna a descrição do material e o status de autorização como JSON
    return jsonify(description=description, authorized=authorized)


# Esta função retorna uma lista de nomes de lojas. É usada para popular  o dropdown e do usuário.
def get_stores():
    stores = []  # Inicializa a lista de lojas

    # Percorre a lista 'lojas' (assumindo que está pré-definida em algum lugar) e adiciona cada nome de loja à lista 'stores'
    for store_name in lojas:
        stores.append(store_name)
    
    return stores  # Retorna a lista de lojas


# Rota principal da aplicação. Apresenta a página inicial e lida com a submissão do formulário para criação de PDF.
@app.route('/', methods=['GET', 'POST'])
def index():
    message = ''  # Mensagem para feedback do usuário
    stores = get_stores()  # Obtém a lista de lojas para exibir na interface

    # Se o método da requisição for POST, processa os dados do formulário
    if request.method == 'POST':
        store = request.form.get('store')  # Pega a loja do formulário
        material = request.form.get('material')  # Pega o material do formulário
        quantity = request.form.get('quantity')  # Pega a quantidade do formulário
        description = request.form.get('description')  # Pega a descrição do formulário

        # Cria o PDF e retorna para o usuário fazer o download
        new_filename = create_pdf(store, material, description, quantity)
        return send_file(new_filename, as_attachment=True, download_name=os.path.basename(new_filename))

    # Retorna a página inicial com as informações necessárias
    return render_template('index.html', message=message, stores=stores, status=status)

# Executa a aplicação
if __name__ == '__main__':
    socketio.run(app, debug=True, host="0.0.0.0", port=5000)
>>>>>>> main
